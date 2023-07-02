# -*- coding: utf-8 -*-
import traceback
import os
import uuid
from collections import defaultdict

import xlrd
import xlwt
from flask import render_template, jsonify, send_from_directory, json, send_file, Response
from flask.ext.babelex import Babel
from flask.ext.celeryext import FlaskCeleryExt
from flask.ext.security import Security, MongoEngineUserDatastore
from flask_admin import Admin
from flask_admin.base import MenuLink
from mongoengine import Q
from werkzeug.utils import secure_filename

from config import APP_ROOT, BACHELORS, SPECIALISTS, RATINGXLS, METHODIC
from modules import app
from modules.accounting.view import AccountingStatisticsView
from modules.address.models import AddressCombinations, District, Region, Settlement
from modules.address.view import AddressView, DistrictView, RegionView, SettlementView
from modules.jiq_connection.token import Token
from modules.profession.models import CertificateEducation, ProgramLearningOutcomes, RegularisationOfProfession
from modules.profession.view import CertificateEducationView, ProgramLearningOutcomesView, KafedraView, \
    RegularisationOfProfessionView
from modules.qualification_type.model import QualificationType
from modules.qualification_type.view import QualificationTypeView
from modules.profession.models import Specialisation, Certificate, CertificateGlobal
from modules.profession.view import CertificateView, CertificateGlobalView
from modules.settings_modules.accounting_statistic.model import SettingAccountingStatistic
from modules.settings_modules.accounting_statistic.view import SettingAccountingStatisticView
from modules.utils.date import StudyYear

ext = FlaskCeleryExt(app)

from modules.document import DocumentType, DocumentTypeView, Order, OrderView
from modules.human import Human
from modules.student import Student, AdditionalScores
from modules.unit.models import AcademicGroup, AbstractUnitLayer, Departament, Kafedra, Unit, GraphicEducationProcess
from modules.unit.views import AcademicGroupView, DepartamentView, InstituteView, UniversityView, PHView, \
    GroupMagAdditionView, GroupBacAdditionView, ArchiveAcademicGroupView, GraduateView, AlienCenterView, \
    VirtualGroupView, ArchiveVirtualGroupView
from modules.profession import Profession, QualificationLevel, ProffOptions, ProffOptionsView, ProfesionView, \
    QualificationLevelView, EducationalProgramMain, SpecialisationView, Qualification, QualificationView, \
    KnowledgeField, \
    KnowledgeFieldView, LevelOfQualification, LevelOfQualificationView, OfficialDurationOfProgramme, \
    OfficialDurationOfProgrammeView, AcquiredCompetences, AcquiredCompetencesView, AcademicRights, AcademicRightsView, \
    AccessRequirements, AccessRequirementsView, EducationalProgramMainView
from modules.subject.view import SubjectView, TeacherView
from modules.subject.models import Subject, Teacher
from modules.views import *

from modules.estimates import Mark, Sheet, SliderView, Slider, SheetView, MarkView, Sabbatical, DiplomaSupplement, \
    SabbaticalView, DiplomaSupplementView, DiplomaProtection, DiplomaProtectionView, StipendCounterPool, EstimateTable
from modules.history import HistoryView, History
from modules.models import Basis, ISCE
from modules.human import HumanView
from modules.student.view import StudentView, StudentHistoryView, DebtorView, \
    PrepareForExpelledStudentsView, LearningView, LearningHistoryView, PrepareForExpelledLearningView, \
    StudentAcademView, LearningAcademView, AdditionalScoresView, MasterView, EdboView, EnrolleeView, \
    DebtorRepeatedCourseView, GraphicEducationProcessView

from modules.import_xls import ImportXLSView, AfterImportView, ExportDataForOrder

from modules.user.form import CustomLoginForm
from modules.user import Role, User, UserView

from modules.permissions.views import RoleView
from modules.tickets.views import TicketView
from modules.tickets.models import Ticket
from modules.forms import TableSizeForm
from modules.generate_mark_file.views import ImportStudentsXml, GenerateMarkFileView, ImportSubjectsView
from modules.type_qualification_works.view import TypesQualificationWorksView
from modules.type_qualification_works.model import TypesQualificationWorks
from modules.global_variables.models import MagazineVariables, Country, GlobalVariables, JuristicPeople, Universities, \
    College
from modules.global_variables.view import MagazineVariablesView, CountryView, GlobalVariablesView, JuristicPeopleView, \
    UniversitiesView, CollegeView
from modules.estimates import RatingStipend
from modules.stipendium import AdditionalStipendiumView
from modules.stipendium import MainStipendiumView
from modules.document.view import RatingPDFView, AccountingView, AppendixToOrderOfGraduation, OrderTransferView, \
    AcademMobilityView
from modules.document import Accounting
from modules.generate_documents.accounting import get_accounting_file

from modules.estimates.models import Enrollee, CompetitiveOffer
from modules.order_edbo.model import OrderEdbo
from modules.order_edbo import OrderEdboView
from modules.record_book.model import RecordBooks
from modules.record_book.view import RecordBookView
from modules.document.models import AcademMobility
from pip._vendor import requests
from modules.estimates.count_estimate_tables import count_estimates_by_slice

user_datastore = MongoEngineUserDatastore(db, User, None)
security = Security(app, user_datastore, login_form=CustomLoginForm)
babel = Babel(app, default_locale='uk', )

celery = ext.celery
from modules.estimates.views import RatingStipendView, StipendOrderView, RatingFixerView, RatingStipendDormitoryView, \
    RatingUnionView, PrevDocEduView, GarantEducView, ImportASServiceView, VirtualSheetView, EstimatesTableView
from modules.student.models import DebtorRepeatedCourse

bed_marks = [u"Не з'явився", u'Недопущений', u'Незараховано']
try:
    import uwsgidecorators

    bed_marks = [u"Не з'явився", u'Недопущений', u'Незараховано']


    @app.errorhandler(Exception)
    def default_handler(e):
        print (traceback.format_exc())
        from config import LOG_PATH
        current_time = datetime.datetime.now().strftime("%Y-%m-%d-%H:%M:%S")
        try:
            user = unicode(current_user)
        except:
            user = 'undefined'

        try:
            referrer = request.referrer
            url = request.path
            get_params = unicode(dict(request.args))
            post_params = unicode(dict(request.form))
            files = unicode(request.files) if request.files else 'No files in query'
            ip = request.remote_addr
        except:
            referrer = 'undefined'
            url = 'undefined'
            get_params = 'undefined'
            post_params = 'undefined'
            files = 'undefined'
            ip = 'undefined'

        with open(os.path.join(LOG_PATH, current_time + '.txt'), 'w') as log:
            log.write(u'Time: {time}\n'.format(time=current_time))
            log.write(u'User: {login}\tIP: {ip}\n'.format(login=user, ip=ip))
            log.write(u'Referrer: {ref}\n'.format(ref=referrer))
            log.write(u'Page: {url}\n'.format(url=url))
            log.write(u'GET: {params}\n'.format(params=get_params))
            log.write(u'POST: {params}\n'.format(params=post_params))
            log.write(u'Files: {params}\n'.format(params=files))
            log.write(traceback.format_exc())
        return render_template('admin/exception_handler.html')


    @uwsgidecorators.cron(-5, -3, -1, -1, -1)
    def crutch_which_fixes_a_bug_elusive(arg):
        for group in AcademicGroup.objects():
            if group.list_students:
                l = set(group.list_students)
                if len(group.list_students) != len(l):
                    group.list_students = list(l)
                    group.save()

    # every first day of month in 01:00 am
    @uwsgidecorators.cron(1, 0, 1, -1, -1)
    def gen_accounting_every_month(arg):
        model = Accounting()
        model.date = datetime.date.today()
        model.formed_by = u'Автоматично'
        model.formed_date = datetime.date.today()
        f = get_accounting_file()
        files = open(f.name, 'rb')
        model.file.new_file()
        model.file.replace(files)
        model.save()
except Exception, e:
    print (e)


@app.context_processor
def inject_global_template_context():
    try:
        messages = Ticket.objects(to=current_user.id, user_read_list__ne=str(current_user.id),
                                  user_delete_list__ne=current_user.id).count()
    except:
        messages = 0
    try:
        current_user.last_activity = datetime.datetime.now()
        current_user.save()
    except:
        pass
    size_form = TableSizeForm()
    if current_user.has_role('Admin') or current_user.has_role('Operator') or current_user.has_role('Moderator') \
            or current_user.has_role('cpo'):
        return dict(new_messages=messages, size_form=size_form, user=1, is_operator=current_user.has_role('Operator'))
    else:
        return dict(new_messages=messages, size_form=size_form, user=0, is_operator=current_user.has_role('Operator'))


admin = Admin(app, u'LoD', translations_path=os.path.dirname(__file__) + '/translations', static_url_path='/static',
              index_view=IndexView(url='/admin'), base_template='admin/base.html')
admin.add_link(MenuLink(u'Вийти', url='/logout'))

# студенти
admin.add_view(StudentView(Student, endpoint='student', name=u'Студенти, що навчаються',
                           category=u'Контингент студентів'))
admin.add_view(StudentAcademView(Student, endpoint='student_academ', name=u'Студенти в академ. відпустці',
                                 academ=True, category=u'Контингент студентів'))
admin.add_view(StudentHistoryView(Student, endpoint='history_student', name=u'Архів студентів',
                                  category=u'Контингент студентів'))
admin.add_view(PrepareForExpelledStudentsView(Student, endpoint='prepare_of_expelled', name=u'Студенти на відрахування',
                                              category=u'Контингент студентів'))

admin.add_view(LearningView(Student, endpoint='learning', name=u'Слухачі, що навчаються',
                            category=u'Контингент слухачів'))
admin.add_view(LearningAcademView(Student, endpoint='learning_academ', name=u'Слухачі в академ. відпустці',
                                  academ=True, category=u'Контингент слухачів'))
admin.add_view(LearningHistoryView(Student, endpoint='history_learning', name=u'Архів слухачів',
                                   category=u'Контингент слухачів'))
admin.add_view(PrepareForExpelledLearningView(Student, endpoint='prepare_of_expelled_learning',
                                              name=u'Слухачі на відрахування',
                                              category=u'Контингент слухачів'))
# Магістри
admin.add_view(MasterView(Student, endpoint='master', name=u'Випускники', category=u'Вступна кампанія'))
admin.add_view(EnrolleeView(Enrollee, endpoint='abit', name=u'Абітурієнти', category=u'Вступна кампанія'))

# групи

admin.add_view(AcademicGroupView(endpoint='group', name=u'Групи', category=u'Групи'))
admin.add_view(ArchiveAcademicGroupView(endpoint='archive_group', name=u'Архів груп', category=u'Групи'))

# віртуальні групи
admin.add_view(
    VirtualGroupView(AcademicGroup, endpoint='virtual_group', name=u'Групи вільного вибору',
                     category=u'Вільний вибір дисциплін'))
admin.add_view(
    ArchiveVirtualGroupView(AcademicGroup, endpoint='archive_virtual_group', name=u'Архів груп вільного вибору',
                            category=u'Вільний вибір дисциплін'))

# успішність
admin.add_view(SabbaticalView(Sabbatical, endpoint='sabbatical', name=u'Академічні довідки', category=u'Успішність'))
admin.add_view(DiplomaSupplementView(DiplomaSupplement, endpoint='diploma_supplement', name=u'Додатки до диплому',
                               category=u'Успішність'))
admin.add_view(SliderView(Slider, endpoint='slider', name=u'Аркуші успішності студента', category=u'Успішність'))
admin.add_view(SheetView(Sheet, endpoint='sheet', name=u'Відомості', category=u'Успішність'))
admin.add_view(
    VirtualSheetView(Sheet, endpoint='virtual_sheet', name=u'Відомості груп вільного вибору', category=u'Успішність'))
admin.add_view(DiplomaProtectionView(DiplomaProtection, endpoint='diploma_protection',
                                     name=u'Захист кваліфікаційних робіт', category=u'Успішність'))
admin.add_view(DebtorView(model=Student, endpoint='debtors', name=u'Боржники', category=u'Успішність'))

admin.add_view(DebtorRepeatedCourseView(model=DebtorRepeatedCourse, endpoint='debtors_repeated_course',
                                        name=u'Боржники, проплата за повторний курс', category=u'Успішність'))

admin.add_view(RatingStipendView(RatingStipend, endpoint='rating_stipend', name=u'Рейтинг', category=u'Успішність'))
admin.add_view(
    RatingStipendDormitoryView(RatingStipend, endpoint='rating_stipend_dormitory', name=u'Рейтинг для гуртожитків',
                               category=u'Успішність'))
admin.add_view(RatingPDFView(endpoint='rating_pdf', name=u'Рейтинги в PDF',
                             category=u'Успішність'))

# Дипломи
# admin.add_view(StudentDiplomaView(Student, endpoint='diploma', name=u'Дипломи'))
# admin.add_view(GradebookView(Student, endpoint='gradebook', name=u'Студентські'))

# структурні підрозділи
admin.add_view(DepartamentView(endpoint='departament', name=u'Факультети', category=u'Структурні підрозділи'))
admin.add_view(InstituteView(endpoint='institute', name=u'Інститути', category=u'Структурні підрозділи'))
admin.add_view(
    GraduateView(endpoint='graduate', name=u'Центр післядипломної освіти', category=u'Структурні підрозділи'))
admin.add_view(
    AlienCenterView(endpoint='alien_center', name=u'Центр міжнародних зв’язків', category=u'Структурні підрозділи'))
admin.add_view(UniversityView(endpoint='university', name=u'Університет', category=u'Структурні підрозділи'))

# навчальний процес
admin.add_view(SubjectView(Subject, endpoint='subject', name=u'Дисципліни', category=u'Навчальний процес'))
admin.add_view(
    ProfesionView(Profession, endpoint='profession', name=u'Спеціальності', category=u'Навчальний процес'))
admin.add_view(
    EducationalProgramMainView(EducationalProgramMain, endpoint='educational_program', name=u'Освітня програма',
                               category=u'Навчальний процес'))
admin.add_view(
    SpecialisationView(Specialisation, endpoint='specialisation', name=u'Спеціалізація',
                       category=u'Навчальний процес'))
admin.add_view(QualificationLevelView(QualificationLevel, endpoint='qualification_level', name=u'Кваліфікаційні рівні',
                                      category=u'Навчальний процес'))
admin.add_view(
    QualificationView(Qualification, endpoint='qualification', name=u'Кваліфікації', category=u'Навчальний процес'))
admin.add_view(
    KnowledgeFieldView(KnowledgeField, endpoint='knowledge_field', name=u'Галузі знань', category=u'Навчальний процес'))
admin.add_view(
    LevelOfQualificationView(LevelOfQualification, endpoint='level_of_qualification', name=u'Рівень кваліфікації',
                             category=u'Навчальний процес'))
admin.add_view(OfficialDurationOfProgrammeView(OfficialDurationOfProgramme, endpoint='official_duration_of_programme',
                                               name=u'Тривалість навчання', category=u'Навчальний процес'))
admin.add_view(
    AcquiredCompetencesView(AcquiredCompetences, endpoint='acquired_competences', name=u'Набуті компетентності',
                            category=u'Навчальний процес'))
admin.add_view(
    ProgramLearningOutcomesView(ProgramLearningOutcomes, endpoint='program_learning',
                                name=u'Програмні результати навчання',
                                category=u'Навчальний процес'))
admin.add_view(AcademicRightsView(AcademicRights, endpoint='academic_rights', name=u'Академічні права',
                                  category=u'Навчальний процес'))
admin.add_view(AccessRequirementsView(AccessRequirements, endpoint='access_requirements', name=u'Вимоги до вступу',
                                      category=u'Навчальний процес'))
admin.add_view(TeacherView(Teacher, endpoint='teachers', name=u'Викладачі', category=u'Навчальний процес'))
admin.add_view(AppendixToOrderOfGraduation(endpoint='appendix', name=u'Додаток до наказу про закінчення навчання',
                                           category=u'Навчальний процес'))
admin.add_view(QualificationTypeView(QualificationType, endpoint='qualification_type', name=u'Ступінь кваліфікацій',
                                     category=u'Навчальний процес'))

admin.add_view(RegularisationOfProfessionView(RegularisationOfProfession, endpoint='regularisation_of_profession',
                                              name=u'Професійна кваліфікація',
                                              category=u'Навчальний процес'))

# накази
admin.add_view(OrderView(Order, endpoint='orders', name=u'Накази'))
admin.add_view(StipendOrderView(u'Наказ (стипендії)', endpoint='stipend_order'))
admin.add_view(OrderTransferView(Order, u'Накази про переведення на курс', endpoint='transfer_order'))

# звітність по контингенту на факультет
admin.add_view(AccountingView(Accounting, endpoint='accounting', name=u'Звітність'))

# Service
admin.add_view(ImportXLSView(name=u'Імпорт XLS', endpoint='import', category=u'Сервіс'))
admin.add_view(AfterImportView(name=u'Післяімпортна обробка', endpoint='next_import', category=u'Сервіс'))
admin.add_view(GenerateMarkFileView(name=u'Експорт даних по оцінкам', endpoint='gen_mark_file', category=u'Сервіс'))
admin.add_view(ImportSubjectsView(name=u'Імпорт дисциплін', endpoint='import_subjects', category=u'Сервіс'))
admin.add_view(ImportStudentsXml(name=u'Імпорт даних з xml', endpoint='import_student_info', category=u'Сервіс'))
admin.add_view(
    ImportASServiceView(name=u'Імпорт додаткових балів з xls', endpoint='import_additional_scores', category=u'Сервіс'))
admin.add_view(ExportDataForOrder(name=u'Експорт даних для наказу', endpoint='export_order_data', category=u'Сервіс'))
admin.add_view(
    PrevDocEduView(name=u'Імпорт попереднього документа про освіту', endpoint='ImportPrevDocEdu', category=u'Сервіс'))
admin.add_view(
    RecordBookView(model=RecordBooks, name=u'Номера залікової книжки', endpoint='record_books', category=u'Сервіс'))
# admin.add_view(MagazineWorkProfessionView(name=u'Журнал робітничої професії',
#                                           endpoint='magazine_work_profession', category=u'Сервіс'))

# Загальні дані
admin.add_view(DocumentTypeView(DocumentType, endpoint='doc_type', name=u'Типи документів', category=u'Загальні дані'))
admin.add_view(AcademMobilityView(AcademMobility, endpoint='academ_mobility', name=u'Академічна мобільність',
                                  category=u'Загальні дані'))
admin.add_view(ReasonView(Reason, endpoint='reason', name=u'Причини відрахування', category=u'Загальні дані'))
admin.add_view(AcademReasonView(AcademReason, endpoint='academreason', name=u'Причини академ. відпустки',
                                category=u'Загальні дані'))
admin.add_view(ContinueReasonView(ContinueReason, endpoint='continuereason',
                                  name=u'Причини продовження терміну складання державної атестації',
                                  category=u'Загальні дані'))
admin.add_view(BasisView(Basis, endpoint='basis', name=u'Підстави відрахування', category=u'Загальні дані'))
admin.add_view(AcademBasisView(AcademBasis, endpoint='academ_basis', name=u'Підстави академ. відпустки',
                               category=u'Загальні дані'))
admin.add_view(GlobalVariablesView(GlobalVariables, name=u'Глобальні дані', endpoint='global_variables',
                                   category=u'Загальні дані'))
admin.add_view(CountryView(Country, name=u'Країни', endpoint='countries', category=u'Загальні дані'))
admin.add_view(TypesQualificationWorksView(model=TypesQualificationWorks, name=u'Типи кваліфікаційних робіт',
                                           endpoint='qualification_works', category=u'Загальні дані'))
admin.add_view(AdditionalScoresView(AdditionalScores, name=u'Додаткові бали', endpoint='additional_scores',
                                    category=u'Загальні дані'))
admin.add_view(
    JuristicPeopleView(JuristicPeople, name=u'Юридичний відділ', endpoint='juristic', category=u'Загальні дані'))
admin.add_view(
    UniversitiesView(Universities, name=u'Вищі навчальні заклади', endpoint='universities', category=u'Загальні дані'))
admin.add_view(
    CollegeView(College, name=u'Заклади освіти (на основі яких відбувався вступ)', endpoint='colleges',
                category=u'Загальні дані'))
admin.add_view(
    AddressView(model=AddressCombinations, endpoint='postal', name=u'Поштові коди', category=u'Загальні дані'))
admin.add_view(DistrictView(model=District, endpoint='district', name=u'Області', category=u'Загальні дані'))
admin.add_view(RegionView(model=Region, endpoint='region', name=u'Райони', category=u'Загальні дані'))
admin.add_view(SettlementView(model=Settlement, endpoint='settlement', name=u'Нас. пункти', category=u'Загальні дані'))
admin.add_view(CertificateView(model=Certificate, endpoint='certificate', name=u'Сертифікати спеціальності',
                               category=u'Загальні дані'))
admin.add_view(CertificateEducationView(model=CertificateEducation, endpoint='certificate_education',
                                        name=u'Сертифікати освітньої програми',
                                        category=u'Загальні дані'))
admin.add_view(CertificateGlobalView(model=CertificateGlobal, endpoint='certificate_global',
                                     name=u'Сертифікати 4 рівня акредитації',
                                     category=u'Загальні дані'))

admin.add_view(
    GarantEducView(model=EducationalProgramMain, endpoint='grant_education', name=u'Гаранти освітньої програми',
                   category=u'Загальні дані'))

admin.add_view(
    KafedraView(model=Kafedra, endpoint='kafedra', name=u'Кафедри',
                category=u'Загальні дані'))
admin.add_view(ISCEView(ISCE, endpoint='isce', name=u'ISCE', category=u'Загальні дані'))

# Адміністрування
admin.add_view(HumanView(Human, endpoint='human', name=u'Особові картки студентів', category=u'Адміністрування'))
admin.add_view(NewsView(model=News, name=u'Новини', endpoint='news', category=u'Адміністрування'))
admin.add_view(UserView(User, name=u'Користувачі', endpoint='user', category=u'Адміністрування'))
admin.add_view(HistoryView(History, name=u'Історія', endpoint='history', category=u'Адміністрування'))
admin.add_view(PHView(name=u'П.Х.', endpoint='abstract', category=u'Адміністрування'))
admin.add_view(MagazineVariablesView(MagazineVariables, name=u'Дані на журнал', endpoint='magazine_variables',
                                     category=u'Адміністрування'))
admin.add_view(RoleView(model=Role, endpoint='role', name=u'Ролі', category=u'Адміністрування', views=admin._views))
admin.add_view(ErrorView(u'Логи помилок', category=u'Адміністрування'))
admin.add_view(ProffOptionsView(ProffOptions, endpoint='p_option', name=u'Проф. опції', category=u'Адміністрування'))
admin.add_view(GroupBacAdditionView(endpoint='gb_option', name=u'Академ. Груп.Бак. опції', category=u'Адміністрування'))
admin.add_view(GroupMagAdditionView(endpoint='gm_option', name=u'Академ. Груп.Маг.опції', category=u'Адміністрування'))
admin.add_view(RatingFixerView(endpoint='fix', name=u'Виправлення рейтингу', category=u'Адміністрування'))
admin.add_view(
    RatingUnionView(endpoint='joinSemester', name=u'Стипендія (обєднання семестрів)', category=u'Адміністрування'))

admin.add_view(SettingAccountingStatisticView(SettingAccountingStatistic, endpoint='setting_accounting',
                                              name=u'Налаштування статистики',
                                              category=u'Адміністрування'))

admin.add_view(GraphicEducationProcessView(GraphicEducationProcess, endpoint='graphic_education_process',
                                           name=u'Графік навчального процесу',
                                           category=u'Навчальний процес'))

# стипендія
admin.add_view(MainStipendiumView(endpoint='main_stipendium', name=u'Осн. стипендія', category=u'Стипендії'))
admin.add_view(
    AdditionalStipendiumView(endpoint='additional_stipendium', name=u'Дод. стипендія', category=u'Стипендії'))

admin.add_view(EstimatesTableView(EstimateTable, endpoint='estimates_table', name=u'Генерація таблиць розподілу оцінок',
                                  category=u'Адміністрування'))

admin.add_view(AccountingStatisticsView(name=u'Визначення кількості стипендіатів', endpoint='accountingstatistics'))

admin.add_view(TicketView(name=u'Пошта', endpoint='ticket'))

admin.add_view(MarkView(Mark, endpoint='mark'))

# ЄДБО
admin.add_view(EdboView(Student, endpoint='edbo', category=u'ЄДБО', name=u'ЄДБО'))
admin.add_view(OrderEdboView(OrderEdbo, endpoint='order_edbo', category=u'ЄДБО', name=u'Наказ на зарахування ЄДБО'))

app.jinja_env.globals['isinstance'] = isinstance
app.jinja_env.globals['enumerate'] = enumerate
app.jinja_env.globals['unicode'] = unicode
app.jinja_env.globals['len'] = len
app.jinja_env.globals['reversed'] = reversed
app.jinja_env.globals['datetime'] = datetime
app.jinja_env.globals['Sheet'] = Sheet
app.jinja_env.globals['Slider'] = Slider
app.jinja_env.globals['Sabbatical'] = Sabbatical


@app.route('/create_token')
def create_token():
    t = Token(date_born=datetime.datetime.now(),
              date_death=datetime.datetime.now() + datetime.timedelta(minutes=1), action='sheet.create',
              owner=current_user.login).save()
    return jsonify({'otvet': t.id})


@app.route('/get_token', methods=['GET', 'POST'])
def get_token():
    if request.args["for"] == "slider":
        token_obj = Token.objects(owner=current_user.login, date_death__gt=datetime.datetime.now(),
                                  action='slider.create')

    if request.args["for"] == "sheet":
        token_obj = Token.objects(owner=current_user.login, date_death__gt=datetime.datetime.now(),
                                  action='sheet.create')
    token = None
    if token_obj:
        for t in token_obj:
            if not t.actived:
                if t.date_death > (datetime.datetime.now() + datetime.timedelta(seconds=20)):
                    token = t
    if token is None:
        if request.args["for"] == "slider":
            token = Token(date_born=datetime.datetime.now(),
                          date_death=datetime.datetime.now() + datetime.timedelta(minutes=1),
                          owner=current_user.login, action='slider.create').save()
        if request.args["for"] == "sheet":
            token = Token(date_born=datetime.datetime.now(),
                          date_death=datetime.datetime.now() + datetime.timedelta(minutes=1),
                          owner=current_user.login, action='sheet.create').save()

    delta = (token.date_born + datetime.timedelta(seconds=60) - datetime.datetime.today()).seconds
    print token.id

    arr = [str(token.id), delta, token.action]
    return jsonify({'otvet': arr})


@app.route('/get_files1')
def download_files():
    return send_file(RATINGXLS, attachment_filename="rating.xls", as_attachment=True)


@app.route('/get_files2')
def download_files2():
    return send_file(METHODIC, attachment_filename="rating.doc", as_attachment=True)


@app.route('/init_rating_date')
def init_rating_date():
    if not current_user.has_role('Admin'):
        return 'not allowed'

    courses = Student.objects.distinct('course')
    print courses
    not_expelled = Student.objects(is_expelled=False)

    for course in courses:
        print course
        filtered = not_expelled.filter(course=course)
        RatingStipend.objects(student__in=filtered, course=course).update(set__date=datetime.datetime(2017, 12, 12))
        RatingStipend.objects(student__in=filtered, course__lt=course).update(set__date=datetime.datetime(2016, 12, 12))

    RatingStipend.objects(student__in=Student.objects(is_expelled=True)).update(
        set__date=datetime.datetime(2016, 12, 12))
    return 'good'


@app.route('/fix_rating_for_course/<int:course>')
def fix_rating(course):
    if not current_user.has_role('Admin'):
        return 'not allowed'

    from modules.setting import CATEGORY_ADD_SCORES
    d = defaultdict(set)
    for st in Student.objects(is_expelled=False, in_academic=False, course=course):
        rat = RatingStipend.objects(student=st).order_by('-semester').first()
        if not rat:
            continue
        semester = rat.semester
        # print '\n\n{} -- {}:'.format(st, st.group)

        scores_credit = 0
        credits = 0
        for mark in Mark.objects(student=st, semester=semester, for_rating=True):
            scores_credit += mark.scores * mark.credit
            credits += mark.credit

        # print 'scores_credit : {}'.format(scores_credit)
        # print 'credits : {}'.format(credits)

        if not scores_credit or not credits:
            print '\n\n{} -- {}\nscores_credit: {}'.format(st, st.group, scores_credit)
            continue

        marks_rating = scores_credit / credits
        # print 'marks_rating : {}'.format(marks_rating)

        add_scores = st.get_add_score(semester=semester)
        sport = add_scores.get(CATEGORY_ADD_SCORES[0][0], 0)
        science = add_scores.get(CATEGORY_ADD_SCORES[1][0], 0)
        social = add_scores.get(CATEGORY_ADD_SCORES[2][0], 0)

        additional_rating = (sport + science + social) * 0.1
        if additional_rating > 10:
            additional_rating = 10

        # print 'additional_rating : {}'.format(additional_rating)

        rating = marks_rating * 0.9 + additional_rating
        if rating > 100:
            rating = 100
        # print 'rating : {}'.format(rating)

        rating_stipend = RatingStipend.objects(student=st, semester=semester).first()
        # print rating_stipend

        if rating != rating_stipend.rating:
            # print '\n\n{} -- {}:'.format(st, st.group)
            # print 'scores_credit : {}'.format(scores_credit)
            # print 'credits : {}'.format(credits)
            # print 'marks_rating : {}'.format(marks_rating)
            # print 'additional_rating : {}'.format(additional_rating)
            # print 'rating : {}'.format(rating)
            # print rating_stipend
            d[st.group.id].add(semester)
    print d, '\n\n'

    from modules.estimates.update_mark_info import update_destiny_by_document
    for group_id in d:
        for semester in d[group_id]:
            print '\n\ngroup: {}, semester: {}'.format(AcademicGroup.objects(id=group_id).first(), semester)
            students = Student.objects(group=group_id, is_expelled=False).distinct('id')
            ratings = RatingStipend.objects(student__in=students, semester=semester)
            print ratings
            ratings.delete()
            print 'ratings deleted'

            Mark.objects(source__in=Sheet.objects(group=group_id, semester=semester, all_done=True)) \
                .update(set__for_rating=False, set__old_scores=0)
            Mark.objects(source__in=Slider.objects(semester=semester, all_done=True, student__in=students)) \
                .update(set__for_rating=False, set__old_scores=0)
            print 'marks done'

            for sheet in Sheet.objects(group=group_id, semester=semester, all_done=True):
                update_destiny_by_document(sheet)
            print 'sheets done'

            for slider in Slider.objects(semester=semester, all_done=True,
                                         student__in=students):
                update_destiny_by_document(slider)
            print 'sliders done'
    return unicode(d)


@app.route('/check_rating_for_course/<int:course>')
def check_rating(course):
    if not current_user.has_role('Admin'):
        return 'not allowed'

    if not course:
        return 'no data'

    from modules.setting import CATEGORY_ADD_SCORES
    for st in Student.objects(is_expelled=False, in_academic=False, course=course):
        rat = RatingStipend.objects(student=st).order_by('-semester').first()
        if not rat:
            continue
        semester = rat.semester
        # print '\n\n{} -- {}:'.format(st, st.group)

        scores_credit = 0
        credits = 0
        for mark in Mark.objects(student=st, semester=semester, for_rating=True):
            scores_credit += mark.scores * mark.credit
            credits += mark.credit

        # print 'scores_credit : {}'.format(scores_credit)
        # print 'credits : {}'.format(credits)

        if not scores_credit or not credits:
            print '\n\n{} -- {}\nscores_credit: {}'.format(st, st.group, scores_credit)
            continue

        marks_rating = scores_credit / credits
        # print 'marks_rating : {}'.format(marks_rating)

        add_scores = st.get_add_score(semester=semester)
        sport = add_scores.get(CATEGORY_ADD_SCORES[0][0], 0)
        science = add_scores.get(CATEGORY_ADD_SCORES[1][0], 0)
        social = add_scores.get(CATEGORY_ADD_SCORES[2][0], 0)

        additional_rating = (sport + science + social) * 0.1
        if additional_rating > 10:
            additional_rating = 10

        # print 'additional_rating : {}'.format(additional_rating)

        rating = marks_rating * 0.9 + additional_rating
        if rating > 100:
            rating = 100
        # print 'rating : {}'.format(rating)

        rating_stipend = RatingStipend.objects(student=st, semester=semester).first()
        # print rating_stipend

        if rating != rating_stipend.rating:
            print '\n\n{} -- {}:'.format(st, st.group)
            print 'scores_credit : {}'.format(scores_credit)
            print 'credits : {}'.format(credits)
            print 'marks_rating : {}'.format(marks_rating)
            print 'additional_rating : {}'.format(additional_rating)
            print 'rating : {}'.format(rating)
            print rating_stipend
    return 'checked'


@app.route('/fix_group_rating/<string:group_id>/<int:semester>')
def fix_group_rating(group_id, semester):
    if not current_user.has_role('Admin'):
        return 'not allowed'

    if not group_id or not semester:
        return 'no data'

    from modules.estimates.update_mark_info import update_destiny_by_document

    print 'group: {}, semester: {}'.format(AcademicGroup.objects(id=group_id).first(), semester)
    students = Student.objects(group=group_id, is_expelled=False).distinct('id')
    RatingStipend.objects(student__in=students, semester=semester).delete()
    print 'ratings deleted'

    Mark.objects(source__in=Sheet.objects(group=group_id, semester=semester, all_done=True)) \
        .update(set__for_rating=False, set__old_scores=0)
    Mark.objects(source__in=Slider.objects(semester=semester, all_done=True, student__in=students)) \
        .update(set__for_rating=False, set__old_scores=0)
    print 'marks done'

    for sheet in Sheet.objects(group=group_id, semester=semester, all_done=True):
        update_destiny_by_document(sheet)
    print 'sheets done'

    for slider in Slider.objects(semester=semester, all_done=True,
                                 student__in=students):
        update_destiny_by_document(slider)
    print 'sliders done'

    return 'success'


@app.route('/get_student_without_ep')
def get_student_without_ep():
    strr = u""
    for st in Student.objects(group__in=AcademicGroup.objects(unit_layer=AbstractUnitLayer.objects(
            departament=Departament.objects(abbreviated=u"ФЕЕЕМ").first()).first()), educational_program=None):
        strr += u"{0}<br>".format(st.__unicode__())
    return strr


@app.route('/get-group', methods=['POST'])
def get_groups():
    groups = AcademicGroup.objects(is_archive=False)
    return jsonify(dict(group_name=[unicode(i.abbreviated) for i in groups], group_id=[unicode(i.pk) for i in groups]))


@app.route('/get-filtered-group', methods=['GET'])
def get_filtered_groups():
    period = request.args.get('period')
    curs = request.args.get('curs')
    facultet = request.args.get('facultet')

    groups = AcademicGroup.objects(is_archive=False, course=curs)
    return jsonify(dict(group_name=[unicode(i.abbreviated) for i in groups], group_id=[unicode(i.pk) for i in groups],
                        group_course=[unicode(i.course) for i in groups]))
    # groups = AcademicGroup.objects(course=).distinct('id')
    # return query.filter(group__in=groups)


@app.route('/print_data_for_order')
def print_data_for_order():
    import thread
    thread.start_new_thread(RatingStipend.print_data_for_order, ())
    return 'counting...'


@app.route('/multiproc_stipend/<int:without_mag>')
def multiproc_stipend(without_mag=0):
    """
    /multiproc_stipend/1 -- without magister 2 course
    /multiproc_stipend/0 -- with magister 2 course
    /multiproc_stipend -- with magister 2 course

    """

    if not current_user.has_role('Admin'):
        return 'not logged in'
    unit_layers = RatingStipend.objects.distinct('unit_layer')
    StipendCounterPool.count_by_unit_layers(unit_layers,
                                            without_mag_second_course_flag=True if without_mag == 1 else False)
    return 'without_mag' if without_mag else 'with_mag'


import openpyxl


@app.route('/get_openpyxl', methods=['GET'])
def get_openpyxl():
    dp = DiplomaProtection.objects()
    first_day = datetime.datetime(2019, 1, 9, 00, 00)
    last_day = datetime.datetime(2019, 2, 1, 00, 00)
    filepath = os.path.join(APP_ROOT, 'modules/static/openpyxl.xlsx')
    wb = openpyxl.Workbook()
    wb.save(filename=filepath)
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.active
    iter_row = 1
    for i in dp:
        for j in i.students:
            if j.date_protect is not None:
                print j.date_protect
                if (j.date_protect.date() > first_day.date()) and (j.date_protect.date() < last_day.date()):
                    ws.cell(row=iter_row, column=1).value = str(j.student.qualification_level)
                    ws.cell(row=iter_row, column=2).value = '№' + str(j.number_protocol)
                    ws.cell(row=iter_row, column=3).value = str(j.date_protect)
                    ws.cell(row=iter_row, column=4).value = str(j.student.human.country)
                    iter_row = iter_row + 1
    wb.save(filepath)
    return send_from_directory(os.path.join(APP_ROOT, 'modules/static/'), 'openpyxl.xlsx', as_attachment=True)


UPLOAD_FOLDER = os.path.join(APP_ROOT, 'modules/static/files')
ALLOWED_EXTENSIONS = set(['txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'])

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/upload', methods=["POST"])
def upload():
    if request.method == 'POST':
        _id = request.args.get('id');
        try:
            diploma_protection = DiplomaProtection.objects.get(id=_id)
            if not diploma_protection:
                return u'Помилка: не знайдено diploma_protection'
        except Exception:
            diploma_protection = None
            return u'Помилка: не знайдено diploma_protection'

        # check if the post request has the file part
        if 'file' not in request.files:
            return redirect(request.url)
        file = request.files['file']

        # if user does not select file, browser also
        # submit a empty part without filename
        if file.filename == '':
            return redirect(request.url)
        if file:
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

            book = xlrd.open_workbook(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            index = 0
            sheet = book.sheet_by_index(index)
            offset_row = 1
            # offset_column = 3

            for i, rows in enumerate(range(sheet.nrows)):
                if i <= offset_row:  # (Optionally) skip headers
                    continue
                row = sheet.row_values(rows)
                _topic = row[2].replace(" ", "")

                for _st in diploma_protection.students:
                    if _st.student.human.full_name == row[0] and _st.topic.replace(" ", "") == _topic:
                        if row[3]:
                            _st.english_topic = row[3]
                diploma_protection.save()
            print request.form['return_url']
            return redirect(request.form['return_url'])
        return u'Ошибка при загрузке файла!'


@app.route('/back_up_fksa', methods=["GET"])
def back_up_fksa():
    custom_request = requests.get('http://192.168.0.26:8090/back_up',
                                  headers={
                                      'content-type': 'application/json;charset=utf-8'
                                  })

    # У разі успіху обробляємо відповідь
    if 200 <= custom_request.status_code < 220:
        data = json.loads(custom_request.text)

        for i in data:
            gr = AcademicGroup.objects.get(id=data[i]['group']['_id']['$oid'])

            print '>Group on server -', gr, data[i]['group']['course'], u'for_transfered_is_ready: {}'.format(
                data[i]['group']['for_transfered_is_ready']), u'year_study: {}'.format(data[i]['group']['year_study'])
            print '>Group on local -', gr, gr.course, u'for_transfered_is_ready: {}'.format(
                gr.for_transfered_is_ready), u'year_study: {}'.format(gr.year_study)

            gr.course = data[i]['group']['course']
            gr.save()

            # for st in data[i]['students']:
            #     student = Student.objects(pk=st.get(st.keys()[0]).get("_id").get("$oid")).first()
            #
            #     print '>Student on server - ', student, u', course -', st.get(st.keys()[0]).get(
            #         "course"), u', transfered -', st.get(st.keys()[0]).get("transfered"), u', not_transfered -', st.get(
            #         st.keys()[0]).get("not_transfered")
            #     print '>Student on local - ', student, u', course -', student.course, u', transfered -', student.transfered, u', not_transfered -', student.not_transfered, student.expelled_info

            # student.transfered = bool(st.get(st.keys()[0]).get("transfered"))
            # student.not_transfered = bool(st.get(st.keys()[0]).get("not_transfered"))

            # if student.transfered:
            #     if (int(student.course) - int(st.get(st.keys()[0]).get("course"))) > 1:
            #         student.course = int(st.get(st.keys()[0]).get("course")) + 1
            #         student.transfered = True
            # else:
            # student.course = st.get(st.keys()[0]).get("course")

            # gr.year_study = data[i]['group']['year_study']

            # print "////////////////////////////////"
            # print '>Student on server - ', student, u', course -', st.get(st.keys()[0]).get(
            #     "course"), u', transfered -', st.get(st.keys()[0]).get("transfered"), u', not_transfered -', st.get(
            #     st.keys()[0]).get("not_transfered")
            # print '>Student on local - ', student, u', course -', student.course, u', transfered -', student.transfered, u', not_transfered -', student.not_transfered, student.expelled_info
            #
            # for no, z in enumerate(student.expelled_info):
            #     if z.number and (u'297' in z.number or u'294' in z.number or u'299' in z.number):
            #         print 'Expelled nakaz detected! Deleting....'
            #         del student.expelled_info[no]
            # student.save()

            # print '*****************************'

        return "OK"
    else:
        return 'ups'


@app.route('/parse_sudent_from_EDEBO', methods=["GET", "POST"])
def parse_sudent_from_EDEBO():
    print "start"
    for i in [BACHELORS, SPECIALISTS]:
        with open(i) as data_file:
            data = json.load(data_file)
        problem_names = []
        for student in data.get("data"):
            full_name_birth_date = student.get("personName")
            full_name_birth_date_splited = full_name_birth_date.rpartition(" ")
            birth_date = datetime.datetime.strptime(full_name_birth_date_splited[2], '%d.%m.%Y')
            full_name = full_name_birth_date_splited[0]
            full_name = full_name.replace("-", " ")
            full_name = full_name.replace("  ", " ")
            full_name = full_name.strip()
            finding_humans = Human.objects(full_name=full_name, birth_date=birth_date)
            if finding_humans:
                for finding_human in finding_humans:
                    finding_human.personId = str(student.get("personId"))
                    finding_human.personCodeU = student.get("personCodeU")
                    finding_human.save()
            else:
                problem_names.append(full_name)
    return "1"


# <----------------------------------------------------->
@app.route('/get_openpyxl_now', methods=['GET'])
def get_openpyxl_now():
    # dp = Enrollee.objects(is_delete=False)

    unit_layer = None
    groups = []
    if current_user.has_role("Admin"):
        for i in current_user.unit_layer:
            if i.for_vstup_kompany and i.end_date is None:
                group = AcademicGroup.objects(unit_layer=i)
                for gr in group:
                    groups.append(gr)
    else:
        for i in current_user.unit_layer:
            if i.for_vstup_kompany and i.end_date is None:
                unit_layer = i
        groups = AcademicGroup.objects(unit_layer=unit_layer)
    all_id_students = []
    for gr in groups:
        for st in gr.list_students:
            all_id_students.append(st.id)

    group = Student.objects(pk__in=all_id_students)

    dp = Enrollee.objects(student__in=group, is_delete=False).order_by('-result_mark')

    filepath = os.path.join(APP_ROOT, 'modules/static/openpyxl_enrollee.xlsx')
    wb = openpyxl.Workbook()
    wb.save(filename=filepath)
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.active
    iter_row = 1
    for i in dp:
        ws.cell(row=iter_row, column=1).value = str(i.student.human)
        ws.cell(row=iter_row, column=2).value = str(i.competitiveOffer)
        ws.cell(row=iter_row, column=3).value = str(i.finance)
        ws.cell(row=iter_row, column=4).value = str(i.personal_file)
        ws.cell(row=iter_row, column=5).value = str(i.status_statements)
        ws.cell(row=iter_row, column=6).value = str(i.sync_edbo)
        ws.cell(row=iter_row, column=7).value = str(i.original_or_reference)
        ws.cell(row=iter_row, column=8).value = str(i.result_mark)

        iter_row = iter_row + 1
    wb.save(filepath)
    return send_from_directory(os.path.join(APP_ROOT, 'modules/static/'), 'openpyxl_enrollee.xlsx', as_attachment=True)


@app.route('/get_finished_students', methods=['GET'])
def get_finished_students():
    # dp = Enrollee.objects(is_delete=False)

    unit_layers = []
    if current_user.has_role("Admin"):
        for i in current_user.unit_layer:
            if not i.for_vstup_kompany and i.end_date is None:
                unit_layers.append(i)
    else:
        return
    for unit_layer in unit_layers:
        print unit_layer, "**********************************"
        all_students = []
        groups = AcademicGroup.objects(unit_layer=unit_layer, course=4,
                                       qualification_level=QualificationLevel.objects.get(name=u'Бакалавр'),
                                       is_archive__ne=True, education_form=1)
        for gr in groups:
            print "#1#", gr
            for st in gr.list_students:
                if st.human.sex == 1:
                    all_students.append(st)
        groups = AcademicGroup.objects(unit_layer=unit_layer, course=2,
                                       qualification_level=QualificationLevel.objects.get(
                                           name=u'Бакалавр (зі скороченим терміном навчання)'), education_form=1,
                                       is_archive__ne=True)
        for gr in groups:
            print "#2#", gr
            for st in gr.list_students:
                if st.human.sex == 1:
                    all_students.append(st)
        groups = AcademicGroup.objects(unit_layer=unit_layer, education_form=2, is_archive__ne=True)
        for gr in groups:
            print "#3#", gr
            for st in gr.list_students:
                if st.human.sex == 1:
                    all_students.append(st)
        filepath = None
        if unit_layer.switch == 1:
            filepath = os.path.join(APP_ROOT, 'modules/static/' + unit_layer.institute.abbreviated + '.xlsx')
        elif unit_layer.switch == 0:
            filepath = os.path.join(APP_ROOT, 'modules/static/' + unit_layer.departament.abbreviated + '.xlsx')
        if filepath:
            wb = openpyxl.Workbook()
            wb.save(filename=filepath)
            wb = openpyxl.load_workbook(filename=filepath)
            ws = wb.active
            iter_row = 1
            for i in all_students:
                ws.cell(row=iter_row, column=1).value = str(iter_row)
                ws.cell(row=iter_row, column=2).value = str(i.human)
                ws.cell(row=iter_row, column=3).value = str(i.human.birth_date)
                ws.cell(row=iter_row, column=4).value = str(i.human.reg_address) if i.human.reg_address else str(
                    i.human.registration_address)
                ws.cell(row=iter_row, column=5).value = str(i.expelled_date)
                iter_row = iter_row + 1
            wb.save(filepath)
    return "ok"
    # return send_from_directory(os.path.join(APP_ROOT, 'modules/static/'), 'openpyxl_enrollee.xlsx', as_attachment=True)


# <----------------------------------------------------->
@app.route('/get_openpyxl_buhg34', methods=['GET'])
def get_openpyxl_buhg34():
    filepath = os.path.join(APP_ROOT, 'modules/static/buhg.xlsx')
    wb = openpyxl.Workbook()
    # wb.create_sheet(title='123', index=None)
    wb.save(filename=filepath)
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.active

    iter_row = 2
    ws.cell(row=1, column=1).value = u"Name"
    ws.cell(row=1, column=2).value = u"Surname"
    ws.cell(row=1, column=3).value = u"Patronymic"
    ws.cell(row=1, column=4).value = u"Birthday"
    ws.cell(row=1, column=5).value = u"Passport serial"
    ws.cell(row=1, column=6).value = u"Passport number"
    ws.cell(row=1, column=7).value = u"ID number"
    ws.cell(row=1, column=8).value = u"Region"
    ws.cell(row=1, column=9).value = u"District"
    ws.cell(row=1, column=10).value = u"Settlement"
    ws.cell(row=1, column=11).value = u"Street"
    ws.cell(row=1, column=12).value = u"House"
    ws.cell(row=1, column=13).value = u"Corpus"
    ws.cell(row=1, column=14).value = u"Apartments"
    ws.cell(row=1, column=15).value = u"training_dir"
    ws.cell(row=1, column=16).value = u"group"
    ws.cell(row=1, column=17).value = u"course"
    ws.cell(row=1, column=18).value = u"finance"
    ws.cell(row=1, column=19).value = u"education_form"
    ws.cell(row=1, column=20).value = u"educational_program"
    ws.cell(row=1, column=21).value = u"unit_layer"
    ws.cell(row=1, column=22).value = u"record_book"
    ws.cell(row=1, column=23).value = u"phone_number"

    grs = AcademicGroup.objects(unit_layer__in=AbstractUnitLayer.objects(for_vstup_kompany=True))

    # groups = AcademicGroup.objects(qualification_level__in=QualificationLevel.objects(name=u'Магістр'))

    students = Student.objects((Q(group__exists=True) & Q(group__not__in=grs)),
                               date_entry__gt=datetime.datetime(day=1, month=8, year=2021),
                               # finance=u"Контракт",
                               is_expelled=False,
                               in_academic=False,
                               is_delete=False)
    for i, st in enumerate(students):
        print st.finance
        ws.cell(row=iter_row, column=1).value = str(st.human.name)
        ws.cell(row=iter_row, column=2).value = str(st.human.surname)
        ws.cell(row=iter_row, column=3).value = str(st.human.patronymic)
        ws.cell(row=iter_row, column=4).value = str(st.human.birth_date)
        ws.cell(row=iter_row, column=5).value = str(st.human.passport_serial)
        ws.cell(row=iter_row, column=6).value = str(st.human.passport_number)
        ws.cell(row=iter_row, column=7).value = str(st.human.identification_code)
        ws.cell(row=iter_row, column=8).value = (str(st.human.reg_address.region) if st.human.reg_address else '')
        ws.cell(row=iter_row, column=9).value = (str(st.human.reg_address.district) if st.human.reg_address else '')
        ws.cell(row=iter_row, column=10).value = (str(st.human.reg_address.settlement) if st.human.reg_address else '')
        ws.cell(row=iter_row, column=11).value = (str(st.human.reg_address.street) if st.human.reg_address else '')
        ws.cell(row=iter_row, column=12).value = (str(st.human.reg_address.house) if st.human.reg_address else '')
        ws.cell(row=iter_row, column=13).value = (str(st.human.reg_address.corpus) if st.human.reg_address else '')
        ws.cell(row=iter_row, column=14).value = (str(st.human.reg_address.apartments) if st.human.reg_address else '')
        ws.cell(row=iter_row, column=15).value = u'{} {}'.format(st.training_dir.name, st.training_dir.code)
        ws.cell(row=iter_row, column=16).value = str(st.group)
        ws.cell(row=iter_row, column=17).value = str(st.course)
        ws.cell(row=iter_row, column=18).value = str(st.finance)
        ws.cell(row=iter_row, column=19).value = (
            str(u'денна' if st.group.education_form == 1 else u'заочна') if st.group else '')
        ws.cell(row=iter_row, column=20).value = str(st.educational_program.name if st.educational_program else '')
        ws.cell(row=iter_row, column=21).value = (str(st.group.unit_layer) if st.group else '')
        ws.cell(row=iter_row, column=22).value = (str(st.record_book) if st.record_book else '')
        ws.cell(row=iter_row, column=23).value = (str(st.human.phone_number) if st.human.phone_number else '')
        # print st.group.unit_layer
        iter_row = iter_row + 1
    wb.save(filepath)

    return send_from_directory(os.path.join(APP_ROOT, 'modules/static/'), 'buhg.xlsx',
                               as_attachment=True)


# <----------------------------------------------------->

@app.route('/get_xlsx', methods=['GET'])
def get_xlsx():
    grs = AcademicGroup.objects(unit_layer__in=AbstractUnitLayer.objects(for_vstup_kompany=True))

    students = Student.objects((Q(group__exists=True) & Q(group__not__in=grs)),
                               date_entry__gte=datetime.datetime(day=1, month=9, year=2020),
                               is_expelled=False,
                               in_academic=False,
                               is_delete=False)

    facultets = [(u'ФБТЕГП', 'fbtegp'), (u'ФМТ', 'fmt'), (u'ФІРЕН', 'firen'), (u'ФМІБ', 'fmib'), (u'ФІТКІ', 'fitki'),
                 (u'ФКСА', 'fksa'), (u'ФЕЕЕМ', 'feeem'), (u'ІнЕБМД', 'inebmd')]

    for facultet in facultets:

        filepath = os.path.join(APP_ROOT, 'modules/static/facultets/{}.xlsx'.format(facultet[1]))

        wb = openpyxl.Workbook()

        wb.save(filename=filepath)
        wb = openpyxl.load_workbook(filename=filepath)
        ws = wb.active

        iter_row = 2
        ws.cell(row=1, column=1).value = u"Факультет/Інститут"
        ws.cell(row=1, column=2).value = u"Група"
        ws.cell(row=1, column=3).value = u"Ім`я"
        ws.cell(row=1, column=4).value = u"Прізвище"
        ws.cell(row=1, column=5).value = u"По батькові"
        ws.cell(row=1, column=6).value = u"Дата народження"
        ws.cell(row=1, column=7).value = u"Серія паспорту"
        ws.cell(row=1, column=8).value = u"Номер паспорту"
        ws.cell(row=1, column=9).value = u"Номер залікової книжки"
        ws.cell(row=1, column=10).value = u"ID in DB"

        if facultet[0] != u'ІнЕБМД':
            for i, st in enumerate(students):
                if st.group.unit_layer.departament:
                    if st.group.unit_layer.departament.abbreviated == facultet[0]:
                        ws.cell(row=iter_row, column=1).value = str(facultet[0])
                        ws.cell(row=iter_row, column=2).value = str(st.group)
                        ws.cell(row=iter_row, column=3).value = str(st.human.name)
                        ws.cell(row=iter_row, column=4).value = str(st.human.surname)
                        ws.cell(row=iter_row, column=5).value = str(st.human.patronymic)
                        ws.cell(row=iter_row, column=6).value = str(st.human.birth_date)
                        ws.cell(row=iter_row, column=7).value = str(st.human.passport_serial)
                        ws.cell(row=iter_row, column=8).value = str(st.human.passport_number)
                        ws.cell(row=iter_row, column=9).value = str(st.record_book)
                        ws.cell(row=iter_row, column=10).value = str(st.id)
                        iter_row = iter_row + 1
                    wb.save(filepath)
        else:
            for i, st in enumerate(students):
                if st.group.unit_layer.institute:
                    if st.group.unit_layer.institute.abbreviated == facultet[0]:
                        ws.cell(row=iter_row, column=1).value = str(facultet[0])
                        ws.cell(row=iter_row, column=2).value = str(st.group)
                        ws.cell(row=iter_row, column=3).value = str(st.human.name)
                        ws.cell(row=iter_row, column=4).value = str(st.human.surname)
                        ws.cell(row=iter_row, column=5).value = str(st.human.patronymic)
                        ws.cell(row=iter_row, column=6).value = str(st.human.birth_date)
                        ws.cell(row=iter_row, column=7).value = str(st.human.passport_serial)
                        ws.cell(row=iter_row, column=8).value = str(st.human.passport_number)
                        ws.cell(row=iter_row, column=9).value = str(st.record_book)
                        ws.cell(row=iter_row, column=10).value = str(st.id)
                        iter_row = iter_row + 1
                    wb.save(filepath)

    return 'OK'


@app.route('/is_graduate_prostavlator', methods=["GET"])
def is_graduate_prostavlator():
    for st in Student.objects(is_expelled=True):
        if st.reason and (u"у зв`язку із закінченням навчання" in st.reason.text):
            st.is_graduate = True
        elif Enrollee.objects(student=st).first():
            st.is_graduate = True
        else:
            st.is_graduate = False
        st.save()
    return "OK"


@app.route('/change_stipend_date', methods=["GET"])
def change_stipend_date():
    for r in RatingStipend.objects(semester=0):
        r.date = StudyYear.start_of_year()
        r.save()
    return "OK"


@app.route('/reset_result', methods=["GET"])
def reset_result():
    for enrolle in Enrollee.objects(is_delete=False):
        enrolle.result_mark = 0
        enrolle.save()
    return 'OK'


@app.route('/reset_new_first_students', methods=['GET'])
def reset_new_first_students():
    new_first_course_date_creation = datetime.datetime.now()
    new_first_course_date_creation = new_first_course_date_creation.replace(year=2018, month=8, day=1, hour=0, minute=0,
                                                                            second=0, microsecond=0)
    print new_first_course_date_creation.year
    for gr in AcademicGroup.objects(for_transfered_is_ready=True, course=1,
                                    date_creation__gte=new_first_course_date_creation):
        gr.for_transfered_is_ready = False
        gr.save()
        print gr
    return "1"


@app.route('/set_transfered_students', methods=['GET'])
def set_transfered_students():
    for gr in AcademicGroup.objects():
        print gr
        abb = gr.abbreviated.split('-')
        if len(abb) > 1:
            abb = abb[1]
        else:
            continue
        if len(abb) > 1:
            if abb[1] == '7':
                check_semester(2, gr.list_students)
            if abb[1] == '6':
                check_semester(3, gr.list_students)
            if abb[1] == '5':
                check_semester(4, gr.list_students)
            if abb[1] == '4':
                check_semester(5, gr.list_students)
    return "1"


def check_semester(course, students):
    course = int(course)
    for st in students:
        if st.course == course and not st.transfered:
            print st, st.course
            st.transfered = True
            st.save()
        elif st.course == course - 1 and st.transfered:
            print st, st.course
            st.transfered = False
            st.save()
    return


@app.route('/fixer', methods=["GET"])
def fixer():
    enrolles = Enrollee.objects(status_statements=u'До наказу', is_delete=False)
    for i in enrolles:
        students = Student.objects(human=i.student.human)
        counter = Enrollee.objects(student__in=students, is_delete=False, status_statements=u'До наказу')
        if len(counter) > 1:
            print len(counter)
            for j in counter:
                print j.student
    return 'OK'


@app.route('/replace_orders', methods=["GET"])
def replace_orders():
    orders = Order.objects()
    allowed_names = ['175-с', '154-с']

    for i in orders:
        if i.number in allowed_names:
            if i.date.year == 2020:
                print 'yes'
                file = open(os.path.join(APP_ROOT, 'modules/static/fixed_order/' + unicode(i.number) + '.pdf'), 'rb')
                i.file.new_file()
                i.file.replace(file)
                i.save()

    return 'OK'


@app.route('/test', methods=["GET"])
def test():
    specialisations = ['121', '122', '123', '124', '125', '131', '132', '133', '274', '275']
    competitiveOffer = CompetitiveOffer.objects(
        specialityCode__in=specialisations,
        qualificationGroupName=u'Магістр',
        personRequestDateStart__gte=datetime.datetime.strptime('2020-07-01', '%Y-%m-%d')
    ).count()
    print competitiveOffer
    return 'OK'


@app.route('/revert/marks', methods=["GET"])
def revert_student():
    return '123'


@app.route('/cs', methods=["GET"])
def cs():
    prepared_student = []
    humans = Human.objects(registration_address__icontains=u'хмельницька')
    print '>хмельницька'
    for human in humans:
        for student in human.students:
            if isinstance(student, Student) and student.is_expelled == False and student.is_delete == False:
                print student
                prepared_student.append(student)

    humans = Human.objects(registration_address__icontains=u'рівненська')
    print '>рівненська'
    for human in humans:
        for student in human.students:
            if isinstance(student, Student) and student.is_expelled == False and student.is_delete == False:
                print student
                prepared_student.append(student)

    humans = Human.objects(reg_address__district=u"Хмельницька")
    print '>хмельницька--2'
    for human in humans:
        for student in human.students:
            if isinstance(student, Student) and student.is_expelled == False and student.is_delete == False:
                print student
                prepared_student.append(student)

    humans = Human.objects(reg_address__district=u"Рівненська")
    print '>рівненська--2'
    for human in humans:
        for student in human.students:
            if isinstance(student, Student) and student.is_expelled == False and student.is_delete == False:
                print student
                prepared_student.append(student)

    print 'all students'
    filepath = os.path.join(APP_ROOT, 'modules/static/statistika/statistika.xlsx')

    wb = openpyxl.Workbook()

    wb.save(filename=filepath)
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.active

    iter_row = 1

    for i, st in enumerate(prepared_student):
        ws.cell(row=iter_row, column=1).value = str(st.human.surname)
        ws.cell(row=iter_row, column=2).value = str(st.human.name)
        ws.cell(row=iter_row, column=3).value = str(st.human.patronymic)
        ws.cell(row=iter_row, column=4).value = str(st.human.birth_date)
        ws.cell(row=iter_row, column=5).value = str(st.course)
        ws.cell(row=iter_row, column=6).value = (
            str(u'денна' if st.group.education_form == 1 else u'заочна') if st.group else '')
        ws.cell(row=iter_row, column=7).value = str(st.training_dir)
        ws.cell(row=iter_row, column=8).value = str(
            st.human.registration_address) if st.human.registration_address else (
            str(st.human.reg_address) if st.human.reg_address else '')

        iter_row = iter_row + 1

    wb.save(filepath)

    return 'OK'


@app.route('/delete/abit', methods=["GET"])
def test1():
    groups = []
    if current_user.has_role("Admin"):
        for i in AbstractUnitLayer.objects(for_vstup_kompany=True, end_date=None):
            if i.for_vstup_kompany:
                group = AcademicGroup.objects(unit_layer=i)
                for gr in group:
                    groups.append(gr)
    for gr in groups:
        for st in gr.list_students:
            st.delete()

    return 'OK'


@app.route('/test2', methods=["GET"])
def test2():
    print (u'SYSTEM: Debtor Student Update')
    for mark in Mark.objects(rating=u'-'):
        mark.rating = u"Не з'явився"
        mark.save()
    bad_students = Mark.objects(
        (Q(rating=u"Не з'явився") or Q(rating=u'Недопущений') or Q(rating=u'Незараховано'))).distinct('student')

    bad_students = Student.objects(pk="59b40733027cbc456e298e38")

    for stud in bad_students:
        # marks = Mark.objects(student=stud, rating=u'-')
        marks = Mark.objects((Q(rating=u"Не з'явився") or Q(rating=u'Недопущений') or Q(rating=u'Незараховано')),
                             Q(student=stud))
        marks = marks.filter(student=stud)

        marks_count = marks.count()

        status = True
        for m in marks:
            status = False
            # if isinstance(m.source, Sheet):
            tmp_marks = Mark.objects(student=stud, type=m.type, subject=m.subject, credit=m.credit, model=m.model,
                                     semester=m.semester)
            for tmp_m in tmp_marks:
                print tmp_m
                print tmp_m.source
                if isinstance(tmp_m.source, Slider) or isinstance(tmp_m.source, Sabbatical):
                    status = True
                    marks_count -= 1

        if marks_count < 1:
            stud.is_debtor = False
        else:
            if not stud.human.major:
                stud.is_debtor = True

        stud.save()

    return 'OK'


@app.route('/clear_debtor/<string:iid>/<int:sem>', methods=['GET', 'POST'])
def clear_debtor(iid, sem):
    model = Student.objects(pk=iid).first()
    result = ""
    # student_list.append(Student.objects.get(id="56670adf027cbc6414f214af"))
    rating = (u"Не з'явився", u'Недопущений', u'Незараховано', u'Не вивчає', u'Індивідуальний план')
    if model:
        if model.human:
            result += model.human.full_name
        if model:
            if sem == -1:
                for mark in Mark.objects(student=model, rating__in=rating):
                    result += "\n" + mark.__unicode__()
                    sheet = Sheet.objects(id=mark.source.id).first()
                    if not sheet:
                        slider = Slider.objects.get(id=mark.source.id, student=model)
                        if slider:
                            result += "\n" + slider.__unicode__()
                            slider.delete()
                            mark.delete()
                    elif mark in sheet.marks:
                        result += "\n" + sheet.__unicode__()
                        sheet.marks.remove(mark)
                        sheet.save()
                    mark.delete()
            else:
                for mark in Mark.objects(student=model, rating__in=rating, semester=sem):
                    result += "\n" + mark.__unicode__()
                    sheet = Sheet.objects(id=mark.source.id).first()
                    if not sheet:
                        slider = Slider.objects.get(id=mark.source.id, student=model)
                        if slider:
                            result += "\n" + slider.__unicode__()
                            slider.delete()
                            mark.delete()
                    elif mark in sheet.marks:
                        result += "\n" + sheet.__unicode__()
                        sheet.marks.remove(mark)
                        sheet.save()
                    mark.delete()
    return result


@app.route('/remake_record_books', methods=['GET', 'POST'])
def remake_record_books():
    grs = AcademicGroup.objects(unit_layer__in=AbstractUnitLayer.objects(for_vstup_kompany=True))

    for gr in grs:
        print ">group:", gr

    students = Student.objects((Q(group__exists=True) & Q(group__not__in=grs)),
                               date_entry=datetime.datetime(day=1, month=9, year=2018),
                               is_expelled=False,
                               in_academic=False,
                               is_delete=False).order_by('group')
    record_start_numbers = {}
    year = str(datetime.datetime.now().year)
    year = year[2] + year[3]
    abs = AbstractUnitLayer.objects(end_date=None, for_vstup_kompany__ne=True)
    i = 1
    for ab in abs:
        if ab.switch == 0:
            if ab.departament and ab.departament.code:
                record_start_numbers[ab.departament.abbreviated] = ab.departament.code + year + '000'
        elif ab.switch == 1:
            if ab.institute and ab.institute.code:
                record_start_numbers[ab.institute.abbreviated] = ab.institute.code + year + '000'
    for j, st in enumerate(students):
        if st.educational_program and st.educational_program.profession:
            st_fac = AbstractUnitLayer.objects(professions=st.educational_program.profession, end_date=None,
                                               switch__lte=1).first()
            if st_fac and st_fac.switch == 0:
                fac = record_start_numbers.get(st_fac.departament.abbreviated)
                if fac and not st.record_book:
                    print st
                    number = int(fac) + 1
                    number = str(number)
                    st.record_book = "0{0}-{1}-{2}".format(number[0], number[1] + number[2],
                                                           number[3] + number[4] + number[5])
                    record_start_numbers[st_fac.departament.abbreviated] = int(number)
                    print st.record_book
            elif st_fac and st_fac.switch == 1:
                fac = record_start_numbers.get(st_fac.institute.abbreviated)
                if fac and not st.record_book:
                    print st
                    number = int(fac) + 1
                    number = str(number)
                    st.record_book = "0{0}-{1}-{2}".format(number[0], number[1] + number[2],
                                                           number[3] + number[4] + number[5])
                    record_start_numbers[st_fac.institute.abbreviated] = int(number)
                    print st.record_book
        else:
            print u"да, такие есть - {}".format(st)
        st.save()
    return "True"


@app.route('/get_average_mark', methods=('GET', 'POST'))
def get_average_mark():
    groups = AcademicGroup.objects(abbreviated__contains=u'КН', course__in=[3], is_archive__ne=True)
    info_groups = list()
    for gr in groups:
        print gr
        info_gr = dict()
        info_gr['abb'] = gr.__unicode__()
        group_students_info = list()
        for st in gr.list_students:
            if st.finance != u'Контракт':
                continue
            print st
            count = 0.0
            sum = 0.0
            for m in Mark.objects(student=st, improved__ne=True, model__ne=2):
                if m.scores:
                    count += 1
                    if m.scores > 89:
                        sum += 5
                    elif m.scores > 74:
                        sum += 4
                    elif m.scores > 59:
                        sum += 3
                    elif m.scores > 34:
                        sum += 2
                    elif m.scores > 0:
                        sum += 1
            result_sum = sum / count if count != 0 else 0
            print result_sum
            student_info = dict()
            student_info['name'] = st.human.full_name
            student_info['average_mark'] = round(result_sum, 3)
            group_students_info.append(student_info)
        info_gr['students_info'] = group_students_info
        info_groups.append(info_gr)
    response = []
    for item in info_groups:
        for sto in item['students_info']:
            response.append(sto)
    print 'aasdasd', response
    sorted_arr = sorted(response, key=lambda k: k['average_mark'])
    sorted_arr = reversed(sorted_arr)

    print response, info_groups

    return render_template('admin/average_mark.html', info_groups=sorted_arr)


@app.route('/get_ponovleniy_students', methods=['GET', 'POST'])
def get_ponovleniy_students():
    date = datetime.datetime.now()
    date.replace(year=2019, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    sts = Student.objects(expelled_info__category=u'Поновлення')
    result = u""
    for st in sts:
        for ex in st.expelled_info:

            if ex.category == u"Поновлення" and ex.date.year == 2019:
                result += st.human.full_name + u": "

                result += ex.category + u" " + ex.date.strftime("%Y/%d/%m") + u"</br>"

    return result


@app.route('/get_statuses/<string:iid>', methods=['GET', 'POST'])
def get_statuses(iid):
    order = Order.objects(id=iid).first()
    for st in order.list_students:
        for tr in st.transfer_order:
            try:
                print st, tr.order.number, tr.order.unit_layer, tr.student_status
            except Exception:
                print st, tr.order, tr.student_status
    return "1"


@app.route('/get_xlsx_without_edbo_pk/', methods=['GET', 'POST'])
def get_xlsx_without_edbo_pk():
    students = Student.objects(edbo_pk=None, in_academic=False, is_graduate=False, is_delete=False,
                               prepare_of_expelled=False, is_expelled=False)
    filepath = os.path.join(APP_ROOT, 'modules/static/without_edbo_pk.xlsx')
    wb = openpyxl.Workbook()
    # wb.create_sheet(title='123', index=None)
    wb.save(filename=filepath)
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.active
    iter_row = 1
    for i, st in enumerate(students):
        ws.cell(row=iter_row, column=1).value = str(st.human.full_name)
        ws.cell(row=iter_row, column=2).value = str(st.uuid)
        iter_row = iter_row + 1
    wb.save(filepath)

    return send_from_directory(os.path.join(APP_ROOT, 'modules/static/'), 'without_edbo_pk.xlsx',
                               as_attachment=True)


@app.route('/set_edbo_pk_from_xlsx/', methods=['GET', 'POST'])
def set_edbo_pk_from_xlsx():
    filepath = os.path.join(APP_ROOT, 'modules/static/with_edbo_pk.xlsx')
    book = xlrd.open_workbook(filepath)
    index = 0
    sheet = book.sheet_by_index(index)
    for i, rows in enumerate(range(sheet.nrows)):

        row = sheet.row_values(rows)
        st = Student.objects(uuid=row[3]).first()
        if st and row[0]:
            st.edbo_pk = row[0]
            st.save()
        else:
            print "some was wrang"

    return "All is ok"


@app.route('/set_numbers_for_subjects', methods=['GET', 'POST'])
def set_numbers_for_subjects():
    number = 1000
    for s in Subject.objects():
        s.number = number
        number += 1
        s.save()
    return "all numbets are setted"


@app.route('/get_students_fitki_bak/', methods=['GET', 'POST'])
def get_students_fitki_bak():
    from modules.generate_documents.diploma_supplement import get_data_marks
    unit_layers = []
    if current_user.has_role("Admin"):
        for i in current_user.unit_layer:
            if not i.for_vstup_kompany and i.end_date is None and i.departament and i.departament.abbreviated == u"ФІТКІ":
                unit_layers.append(i)
    else:
        return
    for unit_layer in unit_layers:
        print unit_layer, "**********************************"
        all_students = []
        groups = AcademicGroup.objects(unit_layer=unit_layer, course=4,
                                       qualification_level=QualificationLevel.objects.get(name=u'Бакалавр'),
                                       is_archive=True, education_form=1, )
        for gr in groups:
            if "16" in gr.abbreviated:
                print "#1#", gr

                for st in Student.objects(expelled_group=gr):
                    all_students.append(st)
        groups = AcademicGroup.objects(unit_layer=unit_layer, course=2,
                                       qualification_level=QualificationLevel.objects.get(
                                           name=u'Бакалавр (зі скороченим терміном навчання)'), education_form=1,
                                       is_archive=True)
        for gr in groups:
            if "18" in gr.abbreviated:
                print "#2#", gr

                for st in Student.objects(expelled_group=gr):
                    all_students.append(st)
        filepath = None
        if unit_layer.switch == 1:
            filepath = os.path.join(APP_ROOT, 'modules/static/' + unit_layer.institute.abbreviated + '.xlsx')
        elif unit_layer.switch == 0:
            filepath = os.path.join(APP_ROOT, 'modules/static/' + unit_layer.departament.abbreviated + '.xlsx')
        if filepath:
            wb = openpyxl.Workbook()
            wb.save(filename=filepath)
            wb = openpyxl.load_workbook(filename=filepath)
            ws = wb.active
            iter_row = 1
            for i in all_students:
                try:
                    ds = DiplomaSupplement.objects(student=i).first()
                    data = get_data_marks(ds)

                    ws.cell(row=iter_row, column=1).value = str(iter_row)
                    ws.cell(row=iter_row, column=2).value = str(i.human)
                    ws.cell(row=iter_row, column=3).value = str(data["average_mark"])
                    for i, mark in enumerate(data.get('marks_dict').get('is_certification'), start=4):
                        ws.cell(row=iter_row, column=i).value = mark.get("scores")
                    iter_row = iter_row + 1
                except:
                    print i
            wb.save(filepath)
    return "ok"


@app.route('/fix_bug_with_groups/', methods=['GET', 'POST'])
def fix_bug_with_groups():
    # суть в обязательных полях для виртуальных групп, которые отсутсвуют у уже созданных обычных (моделька всё таки та же, и бд не дает их редактировать пока эти поля не заполняться)
    # так что просто заполним всем обычным группам эти поля чем попало))))
    for gr in AcademicGroup.objects(is_virtual__ne=True):
        print gr
        gr.virtual_subject = Subject.objects().first()
        gr.virtual_credit = 0
        gr.save()
    return "MAGIC"


@app.route('/set_uuid_for_groups', methods=['GET', 'POST'])
def set_uuid_for_groups():
    info = 0
    for gr in AcademicGroup.objects():
        if not gr.uuid:
            gr.uuid = str(uuid.uuid4())
            gr.save()
            info += 1
    return str(info)


# don`t touch it !!!!!!!!!!!!!!!

@app.route('/fix_virtual_groups', methods=['GET', 'POST'])
def fix_name_virtual_groups():
    for gr in AcademicGroup.objects(is_virtual=True):
        gr.full_name = gr.abbreviated
        if gr.list_students:
            for st in gr.list_students:
                if not gr in st.virtual_group_list:
                    st.virtual_group_list.append(gr)
                    st.save()
        gr.save()
    return "well done"


@app.route('/create_tables', methods=['GET', 'POST'])
def create_tables():
    for e_t in EstimateTable.objects():
        e_t.delete()
    for pr in Profession.objects(with_educational_program=True,
                                 qualification__in=QualificationLevel.objects(name__in=[u"Бакалавр", u"Магістр"])):
        for year in range(2018, 2021):
            if not pr.end_date or (pr.end_date and pr.end_date >= datetime.datetime(year=year, month=12, day=31)):
                e_t = EstimateTable(training_dir=pr, year=year)
                marks = count_estimates_by_slice(e_t.year, e_t.training_dir)
                if marks:
                    for slice in marks:
                        e_t.marks["{0}-{1}".format(slice[0], slice[1])] = marks[slice]
                    e_t.save()
        if not pr.end_date or (pr.end_date and pr.end_date >= datetime.datetime(year=2021, month=6, day=30)):
            e_t = EstimateTable(training_dir=pr, year=2021, type=1)
            marks = count_estimates_by_slice(e_t.year, e_t.training_dir)
            if marks:
                for slice in marks:
                    e_t.marks["{0}-{1}".format(slice[0], slice[1])] = marks[slice]
                e_t.save()
    return "well done"


@app.route('/find_students', methods=['GET'])
def find_students():
    filepath = os.path.join(APP_ROOT, 'modules/static/students_with_truble.xlsx')
    filepath_2 = os.path.join(APP_ROOT, 'modules/static/students_with_truble_next.xlsx')
    book = xlrd.open_workbook(filename=filepath)
    wb = xlwt.Workbook()
    ws = wb.add_sheet('first')
    sheet = book.sheet_by_index(0)
    for i, rows in enumerate(range(sheet.nrows)):
        row = sheet.row_values(rows)
        ws.write(i, 0, row[0])
        ws.write(i, 1, row[1])
        ws.write(i, 2, row[2])
        try:
            st = Student.objects(pk=row[1].strip()).first()
            print st, st.get_group().pk
            ws.write(i, 3, str(st.get_group().pk))
            ws.write(i, 4, st.get_group().__unicode__())
        except Exception:
            pass
    wb.save(filepath_2)
    return "All is ok"


@app.route('/repair_students_groups', methods=['GET'])
def repair_students_groups():
    filepath = os.path.join(APP_ROOT, 'modules/static/students_with_truble_next.xlsx')
    book = xlrd.open_workbook(filename=filepath)
    sheet = book.sheet_by_index(0)
    for i, rows in enumerate(range(sheet.nrows)):
        row = sheet.row_values(rows)
        try:
            st = Student.objects(pk=row[1].strip()).first()
            st.expelled_group = AcademicGroup.objects.get(pk=row[3].strip())
            st.save()
        except Exception as e:
            print st, e
    return "All is ok"


@app.route('/fix_retake_mark/<string:id>', methods=['GET'])
def fix_retake_mark(id):
    # try:
    mark = Mark.objects(id=id).first()
    mark.retake = True
    mark.save()
    return "success, mark.retake = {}".format(mark.retake)
    # except Exception as e:
    #     return e


@app.route('/get_average_diplom_mark', methods=['POST'])
def get_average_diplom_mark():
    from modules.generate_documents.diploma_supplement import get_data_marks
    from modules.generate_documents.diploma_supplement import get_diploma_mark

    try:
        req = request.json
    except:
        return Response(json.dumps({'status': 'fail', 'message': 'request not has json'}), status=406,
                        mimetype='application/json')

    if not req:
        return Response(json.dumps({'status': 'fail', 'message': 'request is empty'}), status=406,
                        mimetype='application/json')
    if 'data' not in req:
        return Response(json.dumps({'status': 'fail', 'message': 'request json not has data'}), status=406,
                        mimetype='application/json')

    data = req['data']
    if not data.get('diploma_series'):
        return Response(json.dumps({'status': 'fail', 'message': 'request json not has diploma series'}), status=400,
                        mimetype='application/json')
    if not data.get('diploma_number'):
        return Response(json.dumps({'status': 'fail', 'message': 'request json not has diploma number'}), status=400,
                        mimetype='application/json')

    student = Student.objects(bachelor_diploma__number="{}.{}".format(data.get('diploma_series'), data.get('diploma_number'))).first()
    if not student:
        return Response(json.dumps({'status': 'fail', 'message': 'there are no students with this data'}), status=400,
                        mimetype='application/json')

    ds = DiplomaSupplement.objects(student=student).first()
    if not ds:
        return Response(json.dumps({'status': 'fail', 'message': 'that student haven`t diploma supplement'}), status=400,
                        mimetype='application/json')

    sheet_happen = False
    try:
        average_mark = get_data_marks(ds).get('average_mark')
        diploma_mark = get_diploma_mark(student).get('scores')
        if not average_mark or not diploma_mark:
            sheet_happen = True
    except:
        sheet_happen = True

    if sheet_happen:
        return Response(json.dumps({'status': 'fail', 'message': 'that student have problems with diploma'}), status=400,
                        mimetype='application/json')

    result = json.dumps({"status": "success", "message": "OK", "average_mark": average_mark,
                         "diploma_mark": diploma_mark})
    status = json.loads(result).get("status")

    return Response(result, status=(200 if status == "success" else 400), mimetype='application/json')


@app.route('/generate_sheet_report/<int:start_year>-<int:end_year>-<string:uuid>-<string:report_type>', methods=['GET'])
def generate_sheet_report_by_subject(start_year, end_year, uuid, report_type):
    from modules.scripts.MarksReports import generate_sheet_report
    return generate_sheet_report(start_year, end_year, uuid, report_type)


@app.route('/transfer_students_to_another_ph/<string:ph_1>-<string:ph_2>', methods=['GET'])
def transfer_students_to_another_ph(ph_1, ph_2):
    ph_1 = AbstractUnitLayer.objects(id=ph_1).first()
    ph_2 = AbstractUnitLayer.objects(id=ph_2).first()
    groups_for_transfer = AcademicGroup.objects(unit_layer=ph_1, is_archive=False)
    for gr in groups_for_transfer:
        gr.unit_layer = ph_2
        gr.save()
    return str(len(groups_for_transfer))


@app.route('/transfer_students_to_another_ph_by_spec/<string:spec>-<string:ph>', methods=['GET'])
def transfer_students_to_another_ph_by_spec(spec, ph):
    spec = Profession.objects(id=spec).first()
    ph = AbstractUnitLayer.objects(id=ph).first()
    groups_for_transfer = AcademicGroup.objects(training_dir=spec, is_archive=False)
    count = 0
    for gr in groups_for_transfer:
        if not gr.unit_layer.for_vstup_kompany and not gr.unit_layer.for_accounting:
            count += 1
            gr.unit_layer = ph
            gr.save()
    return str(count)


@app.route('/get/group_sheets_info/<course>/<semester>', methods=["POST", "GET"])
def group_sheets_info(course, semester):
    from flask import json
    groups = AcademicGroup.objects(course=course, is_virtual=False, is_archive=False)
    result = {}
    for gr in groups:
        gr_name = gr.__unicode__()
        sheets = Sheet.objects(group=gr, course=course, semester=semester)
        if gr.prof_layer:
            if gr.prof_layer.subjects:
                for key in gr.prof_layer.subjects:
                    if gr.prof_layer.subjects[key]["semester"] == int(semester):
                        part_of_key = key.split("@")[0]
                        subj = Subject.objects(id=part_of_key).first()
                        has_sheet = len(sheets.filter(subject=subj))
                        if not has_sheet:
                            if gr_name in result:
                                result[gr_name].append(subj.name)
                            else:
                                result[gr_name] = [subj.name]
    result_string = u""
    for key in result:
        result_string += u"{0}: {1}<br>".format(key, len(result[key]))
        for subj in result[key]:
            result_string += u"----{0}<br>".format(subj)
    return result_string


if __name__ == '__main__':
    host = os.environ.get('LOD_HOST', '0.0.0.0')
    port = int(os.environ.get('LOD_PORT', '5000'))
    app.run(host=host, debug=True, port=port, threaded=True)

    # celery -A server.celery worker
